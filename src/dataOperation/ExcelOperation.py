import openpyxl
import os


class ExcelData:
    global wb
    global excelFilePath
    excelFilePath = os.getcwd() + "\\ExcelData\\data.xlsx"
    wb = openpyxl.load_workbook(excelFilePath)

    # Reading data from Excel file #
    def ReadExcel(sheetName, colName):
        sheet = wb.get_sheet_by_name(sheetName)
        rowNum = sheet.max_row
        ColNum = ExcelData.getColumnName(sheetName, colName)
        for row in range(2, rowNum + 1):
            print(sheet._get_cell(row, ColNum).value)

    # Fetching column number from user input #
    def getColumnName(sheetName, colName):
        columnNumber = 0
        sheet: object = wb.get_sheet_by_name(sheetName)
        colNum = sheet.max_column
        # print(colNum)
        for col in range(1, colNum):
            ColumnName = sheet._get_cell(1, col).value
            if ColumnName == colName:
                columnNumber = col;
                break
        return columnNumber;
